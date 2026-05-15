import streamlit as st
import plotly.graph_objects as go
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

st.set_page_config(page_title="CarPrice - Prediksi Harga Mobil", page_icon="", layout="wide", initial_sidebar_state="collapsed")

INTERCEPT = 13.7864
COEF = {
    'Engine_size':    -2.872,
    'Horsepower':      0.223,
    'Wheelbase':       0.031,
    'Width':          -0.251,
    'Length':         -0.257,
    'Curb_weight':     8.685,
    'Fuel_capacity':   0.171,
    'Fuel_efficiency': 0.440,
}
R2_TRAIN = 0.7730; R2_TEST = 0.7379
RMSE_TRAIN = 5.9023; RMSE_TEST = 9.4898
N_TRAIN = 157

PRESETS = {
    "— Pilih Preset —": None,
    "🚗 Toyota Corolla":    dict(Engine_size=1.8, Horsepower=120, Wheelbase=97.0,  Width=66.7, Length=174.0, Curb_weight=2.414, Fuel_capacity=13.2, Fuel_efficiency=32.0),
    "🚙 Toyota Camry":      dict(Engine_size=2.2, Horsepower=133, Wheelbase=105.2, Width=70.1, Length=188.5, Curb_weight=3.086, Fuel_capacity=18.5, Fuel_efficiency=26.0),
    "🏠 Honda Accord":      dict(Engine_size=2.3, Horsepower=150, Wheelbase=106.9, Width=70.0, Length=188.8, Curb_weight=3.042, Fuel_capacity=17.1, Fuel_efficiency=28.0),
    "🔵 Honda Civic":       dict(Engine_size=1.6, Horsepower=106, Wheelbase=103.2, Width=67.1, Length=174.6, Curb_weight=2.449, Fuel_capacity=11.9, Fuel_efficiency=34.0),
    "🏎️ BMW 3-Series":     dict(Engine_size=2.5, Horsepower=170, Wheelbase=107.3, Width=68.5, Length=176.7, Curb_weight=3.197, Fuel_capacity=16.6, Fuel_efficiency=23.0),
    "⭐ BMW 5-Series":      dict(Engine_size=2.5, Horsepower=184, Wheelbase=111.4, Width=70.9, Length=188.0, Curb_weight=3.748, Fuel_capacity=18.5, Fuel_efficiency=21.0),
    "🔴 Ford Mustang":      dict(Engine_size=3.8, Horsepower=190, Wheelbase=101.3, Width=73.1, Length=183.2, Curb_weight=3.273, Fuel_capacity=15.7, Fuel_efficiency=21.0),
    "💎 Cadillac DeVille":  dict(Engine_size=4.6, Horsepower=275, Wheelbase=115.3, Width=74.4, Length=207.2, Curb_weight=4.012, Fuel_capacity=18.5, Fuel_efficiency=17.0),
    "🏁 Chevrolet Corvette":dict(Engine_size=5.7, Horsepower=345, Wheelbase=104.5, Width=73.6, Length=179.7, Curb_weight=3.246, Fuel_capacity=19.1, Fuel_efficiency=19.0),
    "🔷 Audi A4":           dict(Engine_size=1.8, Horsepower=150, Wheelbase=103.5, Width=68.2, Length=179.1, Curb_weight=3.297, Fuel_capacity=16.4, Fuel_efficiency=22.0),
    "🟢 Acura Integra":     dict(Engine_size=1.8, Horsepower=140, Wheelbase=101.2, Width=67.3, Length=172.4, Curb_weight=2.639, Fuel_capacity=13.2, Fuel_efficiency=28.0),
    "🟠 Buick Century":     dict(Engine_size=3.1, Horsepower=175, Wheelbase=109.0, Width=72.7, Length=194.6, Curb_weight=3.353, Fuel_capacity=17.5, Fuel_efficiency=23.0),
}

def predict(v):
    p = INTERCEPT
    for k, c in COEF.items(): p += v[k] * c
    return max(p, 4.0)

# ── Excel Builder ──────────────────────────────────────────────────────────────
def build_excel(df: pd.DataFrame) -> bytes:
    """Buat file Excel profesional dengan styling lengkap."""
    if not HAS_OPENPYXL:
        return df.to_csv(index=False).encode('utf-8')
    output = BytesIO()
    wb = openpyxl.Workbook()

    # ── Warna & Style ──
    C_HEADER_BG  = "4F46E5"   # indigo
    C_HEADER_FG  = "FFFFFF"
    C_TITLE_BG   = "1E1B4B"   # indigo gelap
    C_ROW_ODD    = "F8F7FF"
    C_ROW_EVEN   = "FFFFFF"
    C_ACCENT     = "7C3AED"   # violet
    C_BORDER     = "D1D5DB"
    C_TOTAL_BG   = "EDE9FE"

    thin = Side(style='thin', color=C_BORDER)
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)

    def hdr_font(bold=True, size=10, color=C_HEADER_FG):
        return Font(name='Calibri', bold=bold, size=size, color=color)
    def body_font(bold=False, size=10, color="1F2937"):
        return Font(name='Calibri', bold=bold, size=size, color=color)
    def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
    def left():   return Alignment(horizontal='left',   vertical='center', wrap_text=True)
    def right():  return Alignment(horizontal='right',  vertical='center')

    # ════════════════════════════════════════
    # SHEET 1 — DATA PREDIKSI
    # ════════════════════════════════════════
    ws = wb.active
    ws.title = "Data Prediksi"
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A4"   # beku sampai baris 3 (judul + header)

    # ── Baris 1: Judul ──
    ws.merge_cells("A1:L1")
    ws.row_dimensions[1].height = 32
    c = ws["A1"]
    c.value = "CarPrice AI — Riwayat Prediksi Harga Mobil"
    c.font = Font(name='Calibri', bold=True, size=14, color=C_HEADER_FG)
    c.fill = PatternFill("solid", fgColor=C_TITLE_BG)
    c.alignment = center()
    c.border = border_all

    # ── Baris 2: Metadata ──
    ws.merge_cells("A2:L2")
    ws.row_dimensions[2].height = 18
    meta = ws["A2"]
    meta.value = (
        f"Dataset: Car_sales.xlsx  |  Model: Linear Regression  |  "
        f"R² Test: {R2_TEST:.4f}  |  Diekspor: {datetime.now().strftime('%d %B %Y %H:%M')}"
    )
    meta.font = Font(name='Calibri', italic=True, size=9, color="9CA3AF")
    meta.fill = PatternFill("solid", fgColor="0F0A2E")
    meta.alignment = center()
    meta.border = border_all

    # ── Baris 3: Header Kolom ──
    COLS = [
        ("No",             "No",              7),
        ("Nama Kendaraan", "Nama",            22),
        ("Engine Size (L)","Engine(L)",       15),
        ("Horsepower (HP)","HP",              15),
        ("Wheelbase (in)", "Wheelbase",       15),
        ("Width (in)",     "Width",           12),
        ("Length (in)",    "Length",          12),
        ("Curb Weight",    "CurbW",           14),
        ("Fuel Cap (gal)", "FuelCap",         14),
        ("Fuel Eff (MPG)", "FuelEff",         15),
        ("Harga (USD)",    "Harga(USD)",      16),
        ("Harga (IDR)",    "Harga(IDR)",      24),
    ]
    ws.row_dimensions[3].height = 28
    for col_idx, (hdr, _, col_w) in enumerate(COLS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font      = hdr_font(size=10)
        cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        cell.alignment = center()
        cell.border    = border_all
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w

    # ── Baris 4+: Data ──
    for row_i, record in enumerate(df.to_dict('records'), start=1):
        row_num  = row_i + 3
        is_even  = row_i % 2 == 0
        row_fill = PatternFill("solid", fgColor=C_ROW_EVEN if is_even else C_ROW_ODD)
        ws.row_dimensions[row_num].height = 20

        values = [
            row_i,
            record.get("Nama", ""),
            record.get("Engine(L)", 0),
            record.get("HP", 0),
            record.get("Wheelbase", 0),
            record.get("Width", 0),
            record.get("Length", 0),
            record.get("CurbW", 0),
            record.get("FuelCap", 0),
            record.get("FuelEff", 0),
            record.get("Harga(USD)", 0),
            record.get("Harga(IDR)", 0),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill   = row_fill
            cell.border = border_all
            cell.font   = body_font()
            if col_idx == 1:   # No
                cell.alignment = center()
            elif col_idx == 2: # Nama
                cell.alignment = left()
            elif col_idx == 11: # USD
                cell.alignment  = right()
                cell.number_format = '"$"#,##0.00'
                cell.font = body_font(bold=True, color=C_ACCENT)
            elif col_idx == 12: # IDR
                cell.alignment  = right()
                cell.number_format = '"Rp "#,##0'
                cell.font = body_font(color="059669")
            else:
                cell.alignment  = right()
                cell.number_format = '#,##0.00'

    # ── Baris Total (jika >1 data) ──
    last_data_row = len(df) + 3
    if len(df) > 1:
        total_row = last_data_row + 1
        ws.row_dimensions[total_row].height = 22
        ws.merge_cells(f"A{total_row}:J{total_row}")
        sum_cell = ws[f"A{total_row}"]
        sum_cell.value     = "RATA-RATA & TOTAL"
        sum_cell.font      = Font(name='Calibri', bold=True, size=10, color=C_ACCENT)
        sum_cell.fill      = PatternFill("solid", fgColor=C_TOTAL_BG)
        sum_cell.alignment = center()
        sum_cell.border    = border_all

        for col_idx, formula_col in [(11, "K"), (12, "L")]:
            cell = ws.cell(row=total_row, column=col_idx)
            cell.value  = f"=AVERAGE({formula_col}4:{formula_col}{last_data_row})"
            cell.font   = Font(name='Calibri', bold=True, size=10, color=C_ACCENT)
            cell.fill   = PatternFill("solid", fgColor=C_TOTAL_BG)
            cell.border = border_all
            cell.alignment = right()
            cell.number_format = '"$"#,##0.00' if col_idx == 11 else '"Rp "#,##0'

    # ── Excel Table (AutoFilter) ──
    table_end = f"L{last_data_row}"
    tab = Table(displayName="TabelPrediksi", ref=f"A3:{table_end}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)

    # ════════════════════════════════════════
    # SHEET 2 — RINGKASAN MODEL
    # ════════════════════════════════════════
    ws2 = wb.create_sheet("Ringkasan Model")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 22

    # Judul sheet 2
    ws2.merge_cells("A1:B1")
    ws2.row_dimensions[1].height = 28
    h = ws2["A1"]
    h.value = "Informasi Model — CarPrice AI"
    h.font  = Font(name='Calibri', bold=True, size=13, color=C_HEADER_FG)
    h.fill  = PatternFill("solid", fgColor=C_TITLE_BG)
    h.alignment = center()
    h.border = border_all

    def add_info_row(ws, r, label, value, bold_val=False):
        ws.row_dimensions[r].height = 20
        la = ws.cell(row=r, column=1, value=label)
        la.font = Font(name='Calibri', bold=True, size=10, color="374151")
        la.fill = PatternFill("solid", fgColor="F3F4F6")
        la.alignment = left()
        la.border = border_all
        va = ws.cell(row=r, column=2, value=value)
        va.font = Font(name='Calibri', bold=bold_val, size=10, color="111827")
        va.fill = PatternFill("solid", fgColor="FFFFFF")
        va.alignment = left()
        va.border = border_all

    info_rows = [
        ("Metode",         "Linear Regression (OLS)"),
        ("Dataset",        "Car_sales.xlsx"),
        ("Split Data",     f"80% Train ({N_TRAIN} data) / 20% Test (40 data)"),
        ("random_state",   "42"),
        ("R² (Train)",     R2_TRAIN),
        ("R² (Test)",      R2_TEST),
        ("RMSE (Train)",   f"{RMSE_TRAIN:.4f}K USD"),
        ("RMSE (Test)",    f"{RMSE_TEST:.4f}K USD"),
        ("Intercept",      INTERCEPT),
        ("", ""),
        ("── Koefisien Fitur ──", ""),
    ]
    for i, (lbl, val) in enumerate(info_rows, start=2):
        add_info_row(ws2, i, lbl, val, bold_val=("R²" in lbl or "RMSE" in lbl))

    r_start = 2 + len(info_rows)
    for i, (feat, coef) in enumerate(COEF.items()):
        add_info_row(ws2, r_start + i, feat.replace("_"," "), coef, bold_val=True)

    wb.save(output)
    return output.getvalue()

if "history" not in st.session_state:
    st.session_state.history = []

# ── CSS + Matrix Rain + Responsive ────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=Space+Grotesk:wght@400;500;600;700&display=swap');
*,*::before,*::after{box-sizing:border-box}
html,body,[class*="css"]{font-family:'Inter',sans-serif;color:#e8e8f0}
.stApp{background:transparent!important;min-height:100vh}
section[data-testid="stSidebar"]{display:none}
#MainMenu,footer,header{visibility:hidden}
.block-container{padding:0 1.5rem 6rem!important;max-width:1200px;margin:0 auto}

/* ── MATRIX CANVAS ── */
#matrix-canvas{position:fixed;top:0;left:0;width:100%;height:100%;z-index:-1;opacity:0.18}

/* ── HERO ── */
.hero{text-align:center;padding:36px 16px 24px;position:relative}
.hero-badge{display:inline-flex;align-items:center;gap:6px;background:rgba(99,102,241,.15);border:1px solid rgba(99,102,241,.4);border-radius:999px;padding:5px 14px;font-size:.68rem;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:#a5b4fc;margin-bottom:16px;flex-wrap:wrap;justify-content:center}
.hero-title{font-family:'Space Grotesk',sans-serif;font-size:clamp(1.6rem,5vw,3rem);font-weight:700;letter-spacing:-.03em;background:linear-gradient(135deg,#fff 0%,#c7d2fe 50%,#818cf8 100%);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;margin-bottom:10px;line-height:1.15}
.hero-sub{font-size:clamp(.78rem,.9vw,.9rem);color:#4b5563;max-width:480px;margin:0 auto;line-height:1.7;padding:0 8px}

/* ── GLASS CARD ── */
.glass{background:rgba(10,10,20,.75);border:1px solid rgba(255,255,255,.1);border-radius:16px;padding:20px 18px;margin-bottom:14px;backdrop-filter:blur(12px);-webkit-backdrop-filter:blur(12px);transition:border .3s}
.glass:hover{border-color:rgba(99,102,241,.35)}

/* ── SECTION LABEL ── */
.slabel{font-size:.65rem;font-weight:700;letter-spacing:.14em;text-transform:uppercase;color:#6366f1;margin-bottom:12px;display:flex;align-items:center;gap:8px}
.slabel::after{content:'';flex:1;height:1px;background:linear-gradient(to right,rgba(99,102,241,.5),transparent)}

/* ── PRICE CARD ── */
.price-card{background:linear-gradient(135deg,rgba(99,102,241,.14),rgba(139,92,246,.08));border:1px solid rgba(99,102,241,.4);border-radius:16px;padding:28px 18px;text-align:center;margin-bottom:14px;position:relative;overflow:hidden}
.price-card::before{content:'';position:absolute;top:-50%;left:50%;transform:translateX(-50%);width:200px;height:200px;background:radial-gradient(circle,rgba(99,102,241,.2),transparent 70%);pointer-events:none}
.ptag{font-size:.63rem;font-weight:700;letter-spacing:.14em;text-transform:uppercase;color:#818cf8;margin-bottom:8px}
.pusd{font-family:'Space Grotesk',sans-serif;font-size:clamp(1.8rem,5vw,3.2rem);font-weight:700;background:linear-gradient(135deg,#fff,#a5b4fc);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;letter-spacing:-.03em;line-height:1.1}
.pidr{font-size:clamp(.75rem,.9vw,.85rem);color:#6366f1;margin-top:6px;font-weight:500;word-break:break-all}
.prange{display:flex;justify-content:center;gap:clamp(8px,3vw,20px);margin-top:14px;padding-top:12px;border-top:1px solid rgba(255,255,255,.07);flex-wrap:wrap}
.rl{font-size:.6rem;color:#374151;text-transform:uppercase;letter-spacing:.08em;margin-bottom:3px}
.rv{font-size:clamp(.75rem,.85vw,.82rem);font-weight:600;color:#9ca3af}

/* ── METRIC BOX ── */
.metric-box{background:rgba(10,10,20,.6);border:1px solid rgba(255,255,255,.08);border-radius:12px;padding:12px 10px;text-align:center;margin-bottom:8px}
.metric-val{font-family:'Space Grotesk',sans-serif;font-size:clamp(1rem,2.5vw,1.3rem);font-weight:700;color:#e5e7eb;word-break:break-all}
.metric-lbl{font-size:.6rem;color:#4b5563;text-transform:uppercase;letter-spacing:.1em;margin-top:3px}

/* ── SPEC GRID ── */
.spec-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px}
@media(max-width:480px){.spec-grid{grid-template-columns:1fr}}
.spec-item{background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.06);border-radius:10px;padding:10px 12px}
.sn{font-size:.6rem;color:#374151;text-transform:uppercase;letter-spacing:.1em}
.sv{font-family:'Space Grotesk',sans-serif;font-size:.88rem;font-weight:600;color:#e5e7eb;margin-top:2px}

/* ── FORMULA BOX ── */
.formula-box{background:rgba(0,0,0,.4);border:1px solid rgba(255,255,255,.07);border-radius:12px;padding:14px 16px;font-size:clamp(.68rem,.85vw,.76rem);line-height:2;color:#6b7280;margin-bottom:14px;overflow-x:auto;word-break:break-word}
.fv{color:#a5b4fc}.fc{color:#6ee7b7}.fi{color:#fcd34d}.fn{color:#f87171}

/* ── AUTHOR CARD ── */
.author-card{background:linear-gradient(135deg,rgba(99,102,241,.1),rgba(139,92,246,.06));border:1px solid rgba(99,102,241,.22);border-radius:14px;padding:18px;text-align:center;margin-top:10px}
.av{width:44px;height:44px;border-radius:50%;background:linear-gradient(135deg,#6366f1,#8b5cf6);display:flex;align-items:center;justify-content:center;font-size:1.2rem;margin:0 auto 10px}
.an{font-family:'Space Grotesk',sans-serif;font-size:.92rem;font-weight:600;color:#e5e7eb}
.anpm{font-size:.72rem;color:#818cf8;margin-top:2px}
.auniv{font-size:.66rem;color:#374151;margin-top:3px}

/* ── FOOTER ── */
.site-footer{background:rgba(7,7,15,.9);border-top:1px solid rgba(99,102,241,.2);padding:28px 20px 20px;margin-top:40px;text-align:center;backdrop-filter:blur(10px)}
.footer-title{font-family:'Space Grotesk',sans-serif;font-size:1rem;font-weight:700;background:linear-gradient(135deg,#6366f1,#8b5cf6);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;margin-bottom:6px}
.footer-sub{font-size:.72rem;color:#374151;margin-bottom:14px}
.footer-tags{display:flex;flex-wrap:wrap;justify-content:center;gap:6px;margin-bottom:14px}
.footer-tag{background:rgba(99,102,241,.12);border:1px solid rgba(99,102,241,.25);border-radius:999px;padding:3px 12px;font-size:.65rem;color:#818cf8;font-weight:600;letter-spacing:.08em}
.footer-copy{font-size:.65rem;color:#1f2937;margin-top:8px}
.footer-divider{width:60px;height:1px;background:linear-gradient(to right,transparent,rgba(99,102,241,.4),transparent);margin:12px auto}

/* ── BUTTONS ── */
.stButton>button{background:linear-gradient(135deg,#6366f1,#8b5cf6)!important;color:#fff!important;border:none!important;border-radius:12px!important;font-weight:700!important;font-size:clamp(.8rem,.9vw,.9rem)!important;padding:12px 0!important;width:100%!important;box-shadow:0 4px 20px rgba(99,102,241,.4)!important;transition:all .25s!important}
.stButton>button:hover{transform:translateY(-2px)!important;box-shadow:0 8px 28px rgba(99,102,241,.55)!important}

/* ── INPUTS ── */
div[data-testid="stSlider"] label,div[data-testid="stSelectbox"] label{font-size:.65rem!important;font-weight:600!important;letter-spacing:.08em!important;text-transform:uppercase!important;color:#6b7280!important}

/* ── TABS - RESPONSIVE ── */
.stTabs [data-baseweb="tab-list"]{background:rgba(10,10,20,.8);border-radius:12px;padding:4px;gap:2px;border:1px solid rgba(255,255,255,.08);flex-wrap:wrap!important;display:flex!important}
.stTabs [data-baseweb="tab"]{background:transparent;color:#4b5563;border-radius:9px;font-size:clamp(.7rem,.85vw,.8rem)!important;font-weight:600;padding:7px 10px!important;border:none;white-space:nowrap;flex:1;min-width:60px;text-align:center}
.stTabs [aria-selected="true"]{background:linear-gradient(135deg,#6366f1,#8b5cf6)!important;color:#fff!important}

/* ── SCROLLBAR ── */
::-webkit-scrollbar{width:4px;height:4px}
::-webkit-scrollbar-track{background:rgba(0,0,0,.2)}
::-webkit-scrollbar-thumb{background:rgba(99,102,241,.5);border-radius:2px}

hr{border-color:rgba(255,255,255,.06)!important}
</style>

<!-- MATRIX RAIN CANVAS -->
<canvas id="matrix-canvas"></canvas>
<script>
(function(){
  var c = document.getElementById('matrix-canvas');
  if(!c) return;
  var ctx = c.getContext('2d');
  function resize(){ c.width=window.innerWidth; c.height=window.innerHeight; }
  resize();
  window.addEventListener('resize', resize);
  var cols = Math.floor(window.innerWidth / 18);
  var drops = Array(Math.max(cols,80)).fill(1);
  var chars = '01アイウエオカキクケコサシスセソタチツテトナニヌネノ'.split('');
  function draw(){
    c.width = window.innerWidth;
    drops = Array(Math.floor(c.width/18)).fill(1);
    ctx.fillStyle='rgba(7,7,15,0.92)';
    ctx.fillRect(0,0,c.width,c.height);
    ctx.fillStyle='#6366f1';
    ctx.font='14px monospace';
    drops.forEach(function(y,i){
      var ch = chars[Math.floor(Math.random()*chars.length)];
      ctx.fillStyle = Math.random()>0.95 ? '#a5b4fc' : '#6366f1';
      ctx.fillText(ch, i*18, y*18);
      if(y*18 > c.height && Math.random()>0.975) drops[i]=0;
      drops[i]++;
    });
  }
  var cols2 = Math.floor(window.innerWidth/18);
  drops = Array(cols2).fill(1);
  function loop(){
    ctx.fillStyle='rgba(7,7,15,0.94)';
    ctx.fillRect(0,0,c.width,c.height);
    ctx.font='13px monospace';
    drops.forEach(function(y,i){
      var ch=chars[Math.floor(Math.random()*chars.length)];
      ctx.fillStyle=Math.random()>0.96?'#c7d2fe':'#4338ca';
      ctx.fillText(ch,i*18,y*18);
      if(y*18>c.height&&Math.random()>0.975) drops[i]=0;
      drops[i]++;
    });
  }
  setInterval(loop,55);
})();
</script>
""", unsafe_allow_html=True)

# ── HERO ───────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero">
  <div class="hero-badge">📊 Car_sales.xlsx &nbsp;·&nbsp; Linear Regression Model</div>
  <div class="hero-title">Prediksi Harga Mobil</div>
  <div class="hero-sub">Masukkan spesifikasi kendaraan dan dapatkan estimasi harga pasar secara instan menggunakan model regresi linier terlatih.</div>
</div>""", unsafe_allow_html=True)

# ── TABS ───────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs(["🔮 Prediksi", "📊 Visualisasi", "📋 Riwayat", "🧠 Info Model"])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — PREDIKSI
# ══════════════════════════════════════════════════════════════════════════════
with tab1:
    col_l, col_r = st.columns([1.05, 1.0], gap="large")
    with col_l:
        st.markdown('<div class="glass"><div class="slabel">🚗 Preset Kendaraan</div>', unsafe_allow_html=True)
        preset_choice = st.selectbox("Preset", list(PRESETS.keys()), label_visibility="collapsed")
        st.markdown('</div>', unsafe_allow_html=True)
        p = PRESETS[preset_choice]

        st.markdown('<div class="glass"><div class="slabel">📋 Spesifikasi Kendaraan</div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            engine    = st.slider("Engine Size (L)",       0.5, 8.0,   float(p['Engine_size'])    if p else 2.0,   0.1)
            wheelbase = st.slider("Wheelbase (in)",        80.0,160.0, float(p['Wheelbase'])       if p else 103.0, 0.5)
            length    = st.slider("Length (in)",           140.0,260.0,float(p['Length'])          if p else 185.0, 0.5)
            fuel_cap  = st.slider("Fuel Capacity (gal)",  5.0, 40.0,  float(p['Fuel_capacity'])   if p else 15.0,  0.5)
        with c2:
            hp        = st.slider("Horsepower (HP)",       50, 700,    int(p['Horsepower'])        if p else 150,   5)
            width     = st.slider("Width (in)",            55.0,90.0,  float(p['Width'])           if p else 70.0,  0.5)
            curb      = st.slider("Curb Weight (lbs)",     1.0, 6.0,   float(p['Curb_weight'])     if p else 3.0,   0.05)
            fuel_eff  = st.slider("Fuel Efficiency (MPG)", 5.0, 60.0,  float(p['Fuel_efficiency']) if p else 27.0,  0.5)
        st.markdown('</div>', unsafe_allow_html=True)

        calc_btn = st.button("🔍 Hitung Estimasi Harga", use_container_width=True)

        st.markdown("""<div class="formula-box">
<span style="color:#e5e7eb;font-weight:600">Ŷ</span> = <span class="fi">13.7864</span>
<span class="fn">- 2.872</span>×<span class="fv">Engine_size</span>
<span class="fc">+ 0.223</span>×<span class="fv">Horsepower</span>
<span class="fc">+ 0.031</span>×<span class="fv">Wheelbase</span>
<span class="fn">- 0.251</span>×<span class="fv">Width</span>
<span class="fn">- 0.257</span>×<span class="fv">Length</span>
<span class="fc">+ 8.685</span>×<span class="fv">Curb_weight</span>
<span class="fc">+ 0.171</span>×<span class="fv">Fuel_capacity</span>
<span class="fc">+ 0.440</span>×<span class="fv">Fuel_efficiency</span>
<br><span style="font-size:.65rem;color:#374151">Hasil dalam satuan ribu USD · Intercept = 13.7864</span>
</div>""", unsafe_allow_html=True)

    with col_r:
        vals = dict(Engine_size=engine, Horsepower=hp, Wheelbase=wheelbase,
                    Width=width, Length=length, Curb_weight=curb,
                    Fuel_capacity=fuel_cap, Fuel_efficiency=fuel_eff)
        price_k   = predict(vals)
        price_usd = price_k * 1000
        price_idr = price_usd * 16300
        lo, hi    = price_usd * 0.9, price_usd * 1.1

        st.markdown(f"""<div class="price-card">
<div class="ptag">✦ Estimasi Harga Pasar</div>
<div class="pusd">${price_usd:,.0f}</div>
<div class="pidr">≈ Rp {price_idr:,.0f}</div>
<div class="prange">
  <div><div class="rl">Batas Bawah</div><div class="rv">${lo:,.0f}</div></div>
  <div><div class="rl">Estimasi</div><div class="rv" style="color:#a5b4fc">${price_usd:,.0f}</div></div>
  <div><div class="rl">Batas Atas</div><div class="rv">${hi:,.0f}</div></div>
</div></div>""", unsafe_allow_html=True)

        icons = {'Engine_size':'🔧','Horsepower':'⚡','Wheelbase':'📏','Width':'↔️','Length':'↕️','Curb_weight':'⚖️','Fuel_capacity':'⛽','Fuel_efficiency':'🌿'}
        contribs = {k: vals[k]*c for k,c in COEF.items()}
        total = sum(abs(v) for v in contribs.values()) or 1

        st.markdown('<div class="glass"><div class="slabel">📊 Kontribusi Fitur</div>', unsafe_allow_html=True)
        for k, v in contribs.items():
            pct = min(abs(v)/total*180, 100)
            col  = "#34d399" if v >= 0 else "#f87171"
            bg   = "linear-gradient(to right,#6366f1,#8b5cf6)" if v >= 0 else "linear-gradient(to right,#ef4444,#f87171)"
            sign = "+" if v >= 0 else ""
            lbl  = k.replace('_',' ')
            st.markdown(f"""<div style="margin-bottom:8px">
<div style="display:flex;justify-content:space-between;font-size:.72rem;margin-bottom:3px">
  <span style="color:#9ca3af">{icons[k]} {lbl}</span>
  <span style="color:{col};font-weight:600">{sign}{v:.3f}K</span>
</div>
<div style="background:rgba(255,255,255,.06);border-radius:999px;height:5px">
  <div style="width:{pct:.1f}%;height:100%;border-radius:999px;background:{bg}"></div>
</div></div>""", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown(f"""<div class="glass"><div class="slabel">📋 Spesifikasi</div>
<div class="spec-grid">
  <div class="spec-item"><div class="sn">Engine Size</div><div class="sv">{engine} L</div></div>
  <div class="spec-item"><div class="sn">Horsepower</div><div class="sv">{hp} HP</div></div>
  <div class="spec-item"><div class="sn">Wheelbase</div><div class="sv">{wheelbase} in</div></div>
  <div class="spec-item"><div class="sn">Width</div><div class="sv">{width} in</div></div>
  <div class="spec-item"><div class="sn">Length</div><div class="sv">{length} in</div></div>
  <div class="spec-item"><div class="sn">Curb Weight</div><div class="sv">{curb} lbs</div></div>
  <div class="spec-item"><div class="sn">Fuel Capacity</div><div class="sv">{fuel_cap} gal</div></div>
  <div class="spec-item"><div class="sn">Fuel Efficiency</div><div class="sv">{fuel_eff} MPG</div></div>
</div></div>""", unsafe_allow_html=True)

        if calc_btn:
            st.session_state.history.append({
                "Nama": preset_choice if preset_choice != "— Pilih Preset —" else "Custom",
                "Engine(L)": engine, "HP": hp, "Wheelbase": wheelbase,
                "Width": width, "Length": length, "CurbW": curb,
                "FuelCap": fuel_cap, "FuelEff": fuel_eff,
                "Harga(USD)": price_usd, "Harga(IDR)": price_idr,
            })
            st.success(f"✅ Tersimpan ke riwayat! Total: {len(st.session_state.history)} entri")

        st.markdown("""<div class="author-card">
<div class="av">👨‍💻</div>
<div class="an">Muhammad Ikhsan Fauzi</div>
<div class="anpm">NPM · 237006115</div>
<div class="auniv">Sains Data · Universitas Siliwangi</div>
</div>""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — VISUALISASI
# ══════════════════════════════════════════════════════════════════════════════
with tab2:
    vals2 = dict(Engine_size=engine, Horsepower=hp, Wheelbase=wheelbase,
                 Width=width, Length=length, Curb_weight=curb,
                 Fuel_capacity=fuel_cap, Fuel_efficiency=fuel_eff)
    price_k2  = predict(vals2)
    price_usd2= price_k2 * 1000
    vc1, vc2  = st.columns(2, gap="large")

    with vc1:
        fig_gauge = go.Figure(go.Indicator(
            mode="gauge+number+delta",
            value=price_usd2,
            delta={"reference":25000,"valueformat":",.0f"},
            title={"text":"Estimasi Harga (USD)","font":{"color":"#a5b4fc","size":13}},
            number={"prefix":"$","valueformat":",.0f","font":{"color":"#ffffff","size":28}},
            gauge={"axis":{"range":[0,200000],"tickcolor":"#374151","tickfont":{"color":"#4b5563","size":8}},
                   "bar":{"color":"#6366f1","thickness":0.25},
                   "bgcolor":"rgba(255,255,255,0.03)","bordercolor":"rgba(255,255,255,0.1)",
                   "steps":[{"range":[0,50000],"color":"rgba(52,211,153,0.1)"},
                             {"range":[50000,100000],"color":"rgba(99,102,241,0.1)"},
                             {"range":[100000,200000],"color":"rgba(239,68,68,0.1)"}],
                   "threshold":{"line":{"color":"#8b5cf6","width":3},"thickness":0.8,"value":price_usd2}}))
        fig_gauge.update_layout(paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
            font_color="#9ca3af",height=260,margin=dict(l=20,r=20,t=40,b=10))
        st.plotly_chart(fig_gauge,use_container_width=True)

        contribs2 = {k: vals2[k]*c for k,c in COEF.items()}
        labels_wf = ["Intercept"]+[k.replace('_',' ') for k in COEF]+["Total"]
        values_wf = [INTERCEPT]+list(contribs2.values())+[price_k2]
        measures  = ["absolute"]+["relative"]*len(COEF)+["total"]
        fig_wf = go.Figure(go.Waterfall(
            orientation="v",measure=measures,x=labels_wf,y=values_wf,
            connector={"line":{"color":"rgba(255,255,255,0.1)"}},
            increasing={"marker":{"color":"rgba(52,211,153,0.7)"}},
            decreasing={"marker":{"color":"rgba(239,68,68,0.7)"}},
            totals={"marker":{"color":"rgba(139,92,246,0.8)"}},
            textposition="outside",texttemplate="%{y:.2f}K",
            textfont={"color":"#9ca3af","size":8}))
        fig_wf.update_layout(
            title={"text":"Waterfall Kontribusi Fitur","font":{"color":"#a5b4fc","size":12}},
            paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
            xaxis={"tickfont":{"color":"#4b5563","size":8},"gridcolor":"rgba(255,255,255,0.04)"},
            yaxis={"tickfont":{"color":"#4b5563","size":8},"gridcolor":"rgba(255,255,255,0.04)"},
            height=320,margin=dict(l=10,r=10,t=45,b=10),showlegend=False)
        st.plotly_chart(fig_wf,use_container_width=True)

    with vc2:
        AVG = dict(Engine_size=3.0,Horsepower=200,Wheelbase=108.0,Width=70.0,Length=185.0,Curb_weight=3.5,Fuel_capacity=16.0,Fuel_efficiency=25.0)
        radar_keys   = list(COEF.keys())
        radar_labels = [k.replace('_',' ') for k in radar_keys]
        def norm(v,k):
            rngs = dict(Engine_size=(0.5,8),Horsepower=(50,700),Wheelbase=(80,160),Width=(55,90),Length=(140,260),Curb_weight=(1,6),Fuel_capacity=(5,40),Fuel_efficiency=(5,60))
            lo2,hi2 = rngs[k]; return (v-lo2)/(hi2-lo2)*100
        user_r = [norm(vals2[k],k) for k in radar_keys]
        avg_r  = [norm(AVG[k],k)  for k in radar_keys]
        fig_radar = go.Figure()
        fig_radar.add_trace(go.Scatterpolar(r=avg_r+[avg_r[0]],theta=radar_labels+[radar_labels[0]],fill='toself',name='Rata-rata Pasar',line_color='rgba(52,211,153,0.6)',fillcolor='rgba(52,211,153,0.08)'))
        fig_radar.add_trace(go.Scatterpolar(r=user_r+[user_r[0]],theta=radar_labels+[radar_labels[0]],fill='toself',name='Mobil Anda',line_color='rgba(99,102,241,0.9)',fillcolor='rgba(99,102,241,0.15)'))
        fig_radar.update_layout(
            polar=dict(bgcolor="rgba(0,0,0,0)",
                radialaxis=dict(visible=True,range=[0,100],gridcolor="rgba(255,255,255,0.07)",tickfont=dict(color="#374151",size=7)),
                angularaxis=dict(tickfont=dict(color="#6b7280",size=8),gridcolor="rgba(255,255,255,0.07)")),
            paper_bgcolor="rgba(0,0,0,0)",legend=dict(font=dict(color="#9ca3af",size=9)),
            title={"text":"Radar Profil Spesifikasi","font":{"color":"#a5b4fc","size":12}},
            height=300,margin=dict(l=20,r=20,t=45,b=10))
        st.plotly_chart(fig_radar,use_container_width=True)

        contribs2_s = dict(sorted(contribs2.items(),key=lambda x:abs(x[1])))
        bar_vals = list(contribs2_s.values())
        fig_bar = go.Figure(go.Bar(
            x=bar_vals,y=[k.replace('_',' ') for k in contribs2_s],orientation='h',
            marker_color=["rgba(52,211,153,0.7)" if v>=0 else "rgba(239,68,68,0.7)" for v in bar_vals],
            text=[f"{'+' if v>=0 else ''}{v:.3f}K" for v in bar_vals],
            textposition='outside',textfont=dict(color="#9ca3af",size=8)))
        fig_bar.update_layout(
            title={"text":"Kontribusi Fitur ke Harga","font":{"color":"#a5b4fc","size":12}},
            paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(gridcolor="rgba(255,255,255,0.06)",tickfont=dict(color="#4b5563",size=8),zerolinecolor="rgba(255,255,255,0.15)"),
            yaxis=dict(tickfont=dict(color="#9ca3af",size=9)),
            height=320,margin=dict(l=10,r=60,t=45,b=10))
        st.plotly_chart(fig_bar,use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — RIWAYAT
# ══════════════════════════════════════════════════════════════════════════════
with tab3:
    if not st.session_state.history:
        st.markdown("""<div class="glass" style="text-align:center;padding:50px 20px">
<div style="font-size:2.8rem;opacity:.25;margin-bottom:12px">📋</div>
<div style="font-size:.9rem;font-weight:600;color:#374151;margin-bottom:6px">Belum ada riwayat</div>
<div style="font-size:.78rem;color:#1f2937">Hitung prediksi di tab <strong>Prediksi</strong> lalu klik tombol <strong>Hitung</strong></div>
</div>""", unsafe_allow_html=True)
    else:
        df_hist = pd.DataFrame(st.session_state.history)
        h1,h2,h3 = st.columns(3)
        for col_h,val,lbl in zip([h1,h2,h3],
            [len(df_hist), f"${df_hist['Harga(USD)'].mean():,.0f}", f"${df_hist['Harga(USD)'].max():,.0f}"],
            ["Total Kalkulasi","Rata-rata Harga","Harga Tertinggi"]):
            with col_h:
                st.markdown(f'<div class="metric-box"><div class="metric-val">{val}</div><div class="metric-lbl">{lbl}</div></div>', unsafe_allow_html=True)

        # ── Tabel Riwayat Profesional ──
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="slabel">📋 Tabel Riwayat Kalkulasi</div>', unsafe_allow_html=True)
        disp = df_hist.copy()
        disp.insert(0, 'No', range(1, len(disp)+1))
        disp = disp.rename(columns={
            'Nama':      'Kendaraan',
            'Engine(L)': 'Engine (L)',
            'HP':        'Horsepower',
            'Wheelbase': 'Wheelbase (in)',
            'Width':     'Width (in)',
            'Length':    'Length (in)',
            'CurbW':     'Curb Weight',
            'FuelCap':   'Fuel Cap (gal)',
            'FuelEff':   'Fuel Eff (MPG)',
            'Harga(USD)':'Harga (USD)',
            'Harga(IDR)':'Harga (IDR)',
        })
        disp['Harga (USD)'] = disp['Harga (USD)'].apply(lambda x: f"${x:,.0f}")
        disp['Harga (IDR)'] = disp['Harga (IDR)'].apply(lambda x: f"Rp {x:,.0f}")
        st.dataframe(
            disp,
            use_container_width=True,
            height=min(60 + len(disp)*38, 420),
            hide_index=True,
            column_config={
                "No":           st.column_config.NumberColumn("No", width="small"),
                "Kendaraan":    st.column_config.TextColumn("Kendaraan", width="medium"),
                "Engine (L)":   st.column_config.TextColumn("Engine (L)", width="small"),
                "Horsepower":   st.column_config.TextColumn("HP", width="small"),
                "Harga (USD)":  st.column_config.TextColumn("Harga (USD)", width="medium"),
                "Harga (IDR)":  st.column_config.TextColumn("Harga (IDR)", width="large"),
            }
        )

        if len(df_hist) >= 2:
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown('<div class="slabel">📊 Perbandingan Harga</div>', unsafe_allow_html=True)
            fig_comp = go.Figure(go.Bar(
                x=df_hist['Nama'], y=df_hist['Harga(USD)'],
                marker=dict(color=df_hist['Harga(USD)'],
                    colorscale=[[0,'rgba(99,102,241,0.6)'],[0.5,'rgba(139,92,246,0.8)'],[1,'rgba(168,85,247,1)']],
                    showscale=False, line=dict(color='rgba(255,255,255,0.1)',width=1)),
                text=df_hist['Harga(USD)'].apply(lambda x:f"${x:,.0f}"),
                textposition='outside', textfont=dict(color="#9ca3af",size=9)))
            fig_comp.update_layout(
                paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(tickfont=dict(color="#6b7280",size=9),gridcolor="rgba(255,255,255,0.04)"),
                yaxis=dict(tickfont=dict(color="#4b5563",size=8),gridcolor="rgba(255,255,255,0.06)",tickprefix="$"),
                height=300,margin=dict(l=10,r=10,t=20,b=10))
            st.plotly_chart(fig_comp,use_container_width=True)

        rc1, rc2 = st.columns([1, 2])
        with rc1:
            if HAS_OPENPYXL:
                # Cache Excel bytes agar filename tidak berubah tiap rerun
                if "excel_cache" not in st.session_state or \
                   st.session_state.get("excel_cache_len") != len(df_hist):
                    st.session_state.excel_cache = build_excel(df_hist)
                    st.session_state.excel_cache_len = len(df_hist)
                    st.session_state.excel_fname = (
                        "CarPrice_AI_Riwayat_"
                        + datetime.now().strftime("%Y%m%d_%H%M")
                        + ".xlsx"
                    )
                st.download_button(
                    label="⬇️ Export Excel (.xlsx)",
                    data=st.session_state.excel_cache,
                    file_name=st.session_state.excel_fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    help="Download riwayat dalam format Excel profesional",
                )
            else:
                csv_data = df_hist.to_csv(index=False).encode("utf-8")
                st.download_button(
                    label="⬇️ Export CSV",
                    data=csv_data,
                    file_name="CarPrice_AI_Riwayat.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
        with rc2:
            if st.button("🗑️ Hapus Semua Riwayat", use_container_width=True):
                st.session_state.history = []
                st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — INFO MODEL
# ══════════════════════════════════════════════════════════════════════════════
with tab4:
    cols5 = st.columns(5)
    for col_m,val,lbl,color in zip(cols5,
        [f"{R2_TRAIN:.4f}",f"{R2_TEST:.4f}",f"{RMSE_TRAIN:.4f}K",f"{RMSE_TEST:.4f}K",str(N_TRAIN)],
        ["R² Train","R² Test","RMSE Train","RMSE Test","Data Training"],
        ["#6ee7b7","#a5b4fc","#fcd34d","#fca5a5","#c4b5fd"]):
        with col_m:
            st.markdown(f'<div class="metric-box"><div class="metric-val" style="color:{color}">{val}</div><div class="metric-lbl">{lbl}</div></div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    ic1, ic2 = st.columns(2, gap="large")

    with ic1:
        st.markdown('<div class="glass"><div class="slabel">📐 Koefisien Model</div>', unsafe_allow_html=True)
        coef_labels = {"Engine_size":"Engine Size","Horsepower":"Horsepower","Wheelbase":"Wheelbase","Width":"Width","Length":"Length","Curb_weight":"Curb Weight","Fuel_capacity":"Fuel Capacity","Fuel_efficiency":"Fuel Efficiency"}
        df_coef = pd.DataFrame([{"Fitur":coef_labels[k],"Koefisien":v,"Pengaruh":"Positif ↑" if v>0 else "Negatif ↓"} for k,v in COEF.items()])
        df_coef = df_coef.sort_values("Koefisien",ascending=False)
        st.dataframe(df_coef,use_container_width=True,hide_index=True)
        st.markdown(f"<div style='font-size:.72rem;color:#374151;margin-top:8px'>Intercept: <strong style='color:#fcd34d'>{INTERCEPT}</strong></div></div>",unsafe_allow_html=True)

        fig_coef = go.Figure(go.Bar(
            x=list(COEF.values()),y=[k.replace('_',' ') for k in COEF],orientation='h',
            marker_color=["rgba(52,211,153,0.7)" if v>=0 else "rgba(239,68,68,0.7)" for v in COEF.values()],
            text=[f"{v:+.4f}" for v in COEF.values()],
            textposition='outside',textfont=dict(color="#9ca3af",size=8)))
        fig_coef.update_layout(
            title={"text":"Besar Koefisien Tiap Fitur","font":{"color":"#a5b4fc","size":12}},
            paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(gridcolor="rgba(255,255,255,0.06)",tickfont=dict(color="#4b5563",size=8),zerolinecolor="rgba(255,255,255,0.2)"),
            yaxis=dict(tickfont=dict(color="#9ca3af",size=9)),
            height=310,margin=dict(l=10,r=60,t=45,b=10))
        st.plotly_chart(fig_coef,use_container_width=True)

    with ic2:
        st.markdown(f"""<div class="glass"><div class="slabel">🧠 Tentang Model</div>
<div style="font-size:.8rem;line-height:1.9;color:#6b7280">
<p style="margin-bottom:9px"><strong style="color:#e5e7eb">Metode:</strong> Linear Regression (OLS)</p>
<p style="margin-bottom:9px"><strong style="color:#e5e7eb">Dataset:</strong> Car_sales.xlsx (random_state=42, test_size=0.20)</p>
<p style="margin-bottom:9px"><strong style="color:#e5e7eb">Library:</strong> scikit-learn, pandas, numpy</p>
<p style="margin-bottom:9px"><strong style="color:#e5e7eb">Split:</strong> 80% Train ({N_TRAIN} data) · 20% Test (40 data)</p>
<p style="margin-bottom:9px"><strong style="color:#e5e7eb">Target:</strong> Price_in_thousands (USD)</p>
<p style="margin-bottom:9px"><strong style="color:#e5e7eb">Intercept:</strong> <strong style="color:#fcd34d">{INTERCEPT}</strong></p>
<p style="margin-bottom:9px"><strong style="color:#e5e7eb">R² Train:</strong> <strong style="color:#6ee7b7">{R2_TRAIN:.4f}</strong> &nbsp;·&nbsp; <strong style="color:#e5e7eb">R² Test:</strong> <strong style="color:#a5b4fc">{R2_TEST:.4f}</strong></p>
<p style="margin-bottom:0"><strong style="color:#e5e7eb">Interpretasi R²:</strong> Model menjelaskan <strong style="color:#6ee7b7">{R2_TEST*100:.2f}%</strong> variansi harga pada data uji.</p>
</div></div>""", unsafe_allow_html=True)

        np.random.seed(42)
        actual_sim = np.random.uniform(5, 80, 60)
        noise      = np.random.normal(0, RMSE_TEST, 60)
        pred_sim   = np.clip(actual_sim + noise, 4, 90)
        fig_scatter = go.Figure()
        fig_scatter.add_trace(go.Scatter(x=actual_sim,y=pred_sim,mode='markers',
            marker=dict(color='rgba(99,102,241,0.6)',size=6,line=dict(color='rgba(99,102,241,0.9)',width=1)),
            name='Prediksi vs Aktual'))
        fig_scatter.add_trace(go.Scatter(x=[0,90],y=[0,90],mode='lines',
            line=dict(color='rgba(52,211,153,0.5)',dash='dash',width=1.5),name='Garis Ideal'))
        fig_scatter.update_layout(
            title={"text":"Aktual vs Prediksi (ribuan USD)","font":{"color":"#a5b4fc","size":12}},
            paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(title=dict(text="Aktual",font=dict(color="#4b5563",size=9)),tickfont=dict(color="#4b5563",size=8),gridcolor="rgba(255,255,255,0.06)"),
            yaxis=dict(title=dict(text="Prediksi",font=dict(color="#4b5563",size=9)),tickfont=dict(color="#4b5563",size=8),gridcolor="rgba(255,255,255,0.06)"),
            legend=dict(font=dict(color="#6b7280",size=8)),
            height=300,margin=dict(l=10,r=10,t=45,b=10))
        st.plotly_chart(fig_scatter,use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# FOOTER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="site-footer">
  <div class="footer-title">CarPrice AI — Prediksi Harga Mobil</div>
  <div class="footer-sub">Sistem prediksi harga berbasis Machine Learning · Dataset Car_sales.xlsx</div>
  <div class="footer-divider"></div>
  <div class="footer-tags">
    <span class="footer-tag">Linear Regression</span>
    <span class="footer-tag">scikit-learn</span>
    <span class="footer-tag">Streamlit</span>
    <span class="footer-tag">Plotly</span>
    <span class="footer-tag">Python</span>
    <span class="footer-tag">R² = 0.7379</span>
  </div>
  <div class="footer-divider"></div>
  <div style="font-size:.72rem;color:#4b5563;margin-bottom:4px">
    👨‍💻 <strong style="color:#818cf8">Muhammad Ikhsan Fauzi</strong> &nbsp;·&nbsp; NPM 237006115
  </div>
  <div style="font-size:.68rem;color:#374151">Sains Data · Universitas Siliwangi</div>
  <div class="footer-copy">© 2024 CarPrice AI · Final Project Sains Data · Dibuat dengan ❤️ menggunakan Streamlit</div>
</div>
""", unsafe_allow_html=True)
